from unicodedata import name
from numpy import double
import openpyxl as opxl
import openpyxl
import math
import datetime
from openpyxl.styles import Font, Border, Side

def indiv_fee(wb:openpyxl.Workbook,fee_dict:dict, member:dict,number:int):
    
    fee_hotel_1 = fee_dict["宿泊費"]
    fee_conference_1 = fee_dict["会議室費"]
    fee_bath_1 = fee_dict["入浴費"]
    fee_bus_all = fee_dict["バス＋高速代"]
    fee_event_1  =fee_dict['イベント費']

    fee_bus_1 = fee_bus_all/number

    grade = member["学年"]
    attend_day = member["参加日数"]
    sleep_day = member["宿泊日数"]
    insurance = member["保険"]

    fee_insurance = 0
    if insurance == 0:
        fee_insurance = 400
    else:
        fee_insurance = 1000

    fee_hotel = fee_hotel_1 * sleep_day
    fee_conference = fee_conference_1 * attend_day
    fee_bath = fee_bath_1 * sleep_day
    fee = 0

    bus_count = member[datetime.datetime(2022, 8, 13, 0, 0)] + member[datetime.datetime(2022, 8, 8, 0, 0)]

    fee_bus = fee_bus_1 * bus_count

    sum = fee_hotel + fee_conference + fee_bath + fee_bus

    if attend_day == 6:
        if grade < 2:
            fee = 47000
            fee_event = 0
        else:
            fee = 60000
            fee_event = fee - fee_hotel - fee_conference - fee_bath - fee_bus
    else:
        if grade < 2:
            fee = sum
            fee_event = 0
        else:
            fee_event = fee_event_1
            fee = sum + fee_event

    if fee_event:
        fee_flag = "あり"
    else:
        fee_flag = "なし"

    fee += fee_insurance

    final_fee = math.ceil(fee/100)*100

    ws = wb.create_sheet(title=member["名前"])
    ws["c1"].value = "サマキャン請求書"
    ws["d2"].value = datetime.date.today()
    ws["d3"].value = member["名前"]+" 様"
    ws["a5"].value = "内訳"
    ws["b5"].value = "金額(円)"
    ws["c5"].value = "詳細"
    ws["d5"].value = "備考"
    ws["a6"].value = "個別費"
    ws["b6"].value = None
    ws["c6"].value = None
    ws["d6"].value = None
    ws["a7"].value = "宿泊費"
    ws["b7"].value = fee_hotel
    ws["c7"].value = f"7150円/泊 x {sleep_day}"
    ws["d7"].value = "3食含まれています"
    ws["a8"].value = "会議室費"
    ws["b8"].value = fee_conference
    ws["c8"].value = f"300円/日 x {attend_day}"
    ws["d8"].value = "基本的な活動場所です"
    ws["a9"].value = "入浴税"
    ws["b9"].value = fee_bath
    ws["c9"].value = f"150円/回 x {sleep_day}"
    ws["d9"].value = "税金です"
    ws["a10"].value = "団体費"
    ws["b10"].value = None
    ws["c10"].value = None
    ws["d10"].value = None
    ws["a11"].value = "バス代"
    ws["b11"].value = int(fee_bus)
    ws["c11"].value = f"{int(fee_bus_1)}円/回 x {bus_count}"
    ws["d11"].value = "厳密には、バス＋高速代です。"
    ws["a12"].value = "その他"
    ws["b12"].value = None
    ws["c12"].value = None
    ws["d12"].value = None
    ws["a13"].value = "イベント費"
    ws["b13"].value = fee_event
    ws["c13"].value = "イベントに用いられる"
    ws["d13"].value = f"{grade}年なので徴収{fee_flag}"
    ws["a14"].value = "合計請求額"
    ws["b14"].value = final_fee
    ws["c14"].value = None
    ws["d14"].value = None
    ws["b17"].value = "この合計金額を7/12までに振り込みお願いします。"
    ws["b20"].value = "上智大学英語研究会　プレジデント"
    ws["d20"].value = "三好 航平"

    line = Side(style="thin", color="000000")
    big_line = Side(style="medium",color="000000")
    double_line = Side(style="double",color="000000")
    
    ws["c1"].font = Font(size="18")
    ws["c1"].border = Border(bottom=line)

    for rows in ws["A5:D5"]:
        for cell in rows:
            cell.border = Border(left=line,right=line,top=line,bottom=double_line)
    for rows in ws["A6:D6"]:
        for cell in rows:
            cell.border = Border(left=line,right=line,bottom=line)
    for rows in ws["A7:D14"]:
        for cell in rows:
            cell.border = Border(left=line,right=line,top=line,bottom=line)
    for rows in ws["B20:D20"]:
        for cell in rows:
            cell.border = Border(bottom=line)
    
    
    ws.column_dimensions['A'].width = 11
    ws.column_dimensions['C'].width = 29
    ws.column_dimensions['D'].width = 27
    return final_fee